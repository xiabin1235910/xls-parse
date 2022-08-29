#!/usr/bin/python

from curses import initscr
from openpyxl import load_workbook

def digitsFor6(number):
    initStr = str(number)

    for x in range(len(initStr), 6):
        initStr = '0' + initStr
    
    return initStr


wb_read = load_workbook(filename = 'src/source/relationship.xlsx')

sheet_ranges = wb_read['relation']

bufferDict = {}

for row in sheet_ranges.iter_rows(min_row=2, max_col=2, max_row=16836):
    newKey = digitsFor6(row[0].value)
    if newKey in bufferDict:
        if row[1].value in bufferDict[newKey]['bufferSet']:
            continue
        else:
            bufferDict[newKey]['bufferSet'].add(row[1].value)
            bufferDict[newKey]['accStr'] += ('; ' + row[1].value)
    else:
        bufferDict[newKey] = {
            'bufferSet': {row[1].value},
            'accStr': row[1].value
        }
        

# print(bufferDict)


# wb_read_tmt = load_workbook(filename = 'src/source/temp.xlsx')

wb_read_tmt = load_workbook(filename = 'src/source/tmt.xlsx')

sheet_ranges_tmt = wb_read_tmt['sheet1']

import sqlite3

conn = sqlite3.connect('test.db')
c = conn.cursor()
print ("数据库打开成功")

c.execute('''CREATE TABLE COMPANY
       (name          TEXT    NOT NULL,
       age            INT     NOT NULL,
       group_code     INT     NOT NULL,
       company_code   TEXT    NOT NULL,
       year           TEXT    NOT NULL,
       part           TEXT    NOT NULL);''')
print ("数据表创建成功")

def yearPart(year, age):
    diff = int(year.split('-')[0]) - age
    if diff >= 1930 and diff <= 1939:
        return '30'
    elif diff >= 1940 and diff <= 1949:
        return '40'
    elif diff >= 1950 and diff <= 1959:
        return '50'
    elif diff >= 1960 and diff <= 1969:
        return '60'
    elif diff >= 1970 and diff <= 1979:
        return '70'
    elif diff >= 1980 and diff <= 1989:
        return '80'
    elif diff >= 1990 and diff <= 1999:
        return '90'
    else:
        return 'unknown'


for row in sheet_ranges_tmt.iter_rows(min_row=2, max_col=6, max_row=988297):
# for row in sheet_ranges_tmt.iter_rows(min_row=2, max_col=6, max_row=389):
    if row[4].value is None:
        continue
    else:
        if row[0].value in bufferDict:
            if row[2].value in bufferDict[row[0].value]['accStr']:
                c.execute("INSERT INTO COMPANY (name, age, group_code, company_code, year, part) \
      VALUES (?, ?, ?, ?, ?, ?)", (row[2].value, row[4].value, 1, row[0].value, row[1].value, yearPart(row[1].value, row[4].value)))
                # print(row[0].value, row[1].value,row[2].value,row[3].value,row[4].value)
            else:
                c.execute("INSERT INTO COMPANY (name, age, group_code, company_code, year, part) \
      VALUES (?, ?, ?, ?, ?, ?)", (row[2].value, row[4].value, 0, row[0].value, row[1].value, yearPart(row[1].value, row[4].value)))
                # print(0, row[0].value, row[1].value,row[2].value,row[3].value,row[4].value)
        else:
            c.execute("INSERT INTO COMPANY (name, age, group_code, company_code, year, part) \
      VALUES (?, ?, ?, ?, ?, ?)", (row[2].value, row[4].value, 0, row[0].value, row[1].value, yearPart(row[1].value, row[4].value)))

conn.commit()
print ("数据插入成功")
conn.close()